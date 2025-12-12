import pandas as pd
from openpyxl.styles import Border, Side
from openpyxl import load_workbook

# قراءة ملف Excel
df = pd.read_excel("lab5.xlsx", sheet_name="Sheet2", header=None)

# مثال: تنظيف عمود معيّن بدون حذفه
df[5] = None

# حفظ الملف مؤقتاً
temp_path = "output.xlsx"
df.to_excel(temp_path, index=False, header=False)

# فتح الملف مرة ثانية لإضافة الحدود
wb = load_workbook(temp_path)
ws = wb.active

# إعداد شكل الحدود
thin = Side(border_style="thin", color="0000FF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# إضافة الحدود لكل الخلايا ما عدا B1:C2
for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                        min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.coordinate not in ["B1", "B2", "C1", "C2"]:
            cell.border = border

# حفظ التغييرات
wb.save("output_with_borders.xlsx")
