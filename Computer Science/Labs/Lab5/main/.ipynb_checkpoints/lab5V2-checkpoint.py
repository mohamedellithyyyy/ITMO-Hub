{
 "cells": [
  {
   "cell_type": "code",
   "execution_count": 34,
   "id": "5df135a2-972e-418d-9be3-f3518c5eb2c3",
   "metadata": {},
   "outputs": [],
   "source": [
    "import pandas as pd\n",
    "from openpyxl.styles import Border, Side\n",
    "from openpyxl import load_workbook"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 35,
   "id": "e5269dfa-fe28-458d-868b-84b88bd0e649",
   "metadata": {},
   "outputs": [],
   "source": [
    "# Read file\n",
    "df = pd.read_excel(\"lab5.xlsx\", sheet_name=\"Sheet2\", header=None)"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 36,
   "id": "6adee1d7-0896-4ba3-910f-6619c1615c32",
   "metadata": {},
   "outputs": [],
   "source": [
    "# Clean column instead of deleting\n",
    "# df[5] = None\n",
    "df = df.drop(df.columns[5], axis=1)"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 37,
   "id": "cc49b41e-d267-4b10-aa59-8122bea2fe14",
   "metadata": {
    "panel-layout": {
     "height": 0,
     "visible": true,
     "width": 100
    }
   },
   "outputs": [],
   "source": [
    "# Save temporary\n",
    "temp_path = \"../output/output.xlsx\"\n",
    "df.to_excel(temp_path, index=False, header=False)"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 38,
   "id": "7ecc5f48-4e88-4a35-8e64-c305539ee578",
   "metadata": {},
   "outputs": [],
   "source": [
    "# Reopen to apply borders\n",
    "wb = load_workbook(temp_path)\n",
    "ws = wb.active"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 2,
   "id": "e8f2ed47-4554-45c2-ad60-3f2e0ed5ac9c",
   "metadata": {},
   "outputs": [
    {
     "ename": "NameError",
     "evalue": "name 'Side' is not defined",
     "output_type": "error",
     "traceback": [
      "\u001b[0;31m---------------------------------------------------------------------------\u001b[0m",
      "\u001b[0;31mNameError\u001b[0m                                 Traceback (most recent call last)",
      "Cell \u001b[0;32mIn[2], line 2\u001b[0m\n\u001b[1;32m      1\u001b[0m \u001b[38;5;66;03m# Blue border\u001b[39;00m\n\u001b[0;32m----> 2\u001b[0m blue \u001b[38;5;241m=\u001b[39m Side(border_style\u001b[38;5;241m=\u001b[39m\u001b[38;5;124m\"\u001b[39m\u001b[38;5;124mthin\u001b[39m\u001b[38;5;124m\"\u001b[39m, color\u001b[38;5;241m=\u001b[39m\u001b[38;5;124m\"\u001b[39m\u001b[38;5;124m0000FF\u001b[39m\u001b[38;5;124m\"\u001b[39m)\n\u001b[1;32m      3\u001b[0m border \u001b[38;5;241m=\u001b[39m Border(left\u001b[38;5;241m=\u001b[39mblue, right\u001b[38;5;241m=\u001b[39mblue, top\u001b[38;5;241m=\u001b[39mblue, bottom\u001b[38;5;241m=\u001b[39mblue)\n\u001b[1;32m      5\u001b[0m \u001b[38;5;66;03m# Apply border only to A3:Y15\u001b[39;00m\n",
      "\u001b[0;31mNameError\u001b[0m: name 'Side' is not defined"
     ]
    }
   ],
   "source": [
    "# Blue border\n",
    "blue = Side(border_style=\"thin\", color=\"0000FF\")\n",
    "border = Border(left=blue, right=blue, top=blue, bottom=blue)\n",
    "\n",
    "# Apply border only to A3:Y15\n",
    "for row in ws.iter_rows(min_row=3, max_row=15, min_col=1, max_col=25):  # A=1, Y=25\n",
    "    for cell in row:\n",
    "        cell.border = border\n"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 41,
   "id": "e602ebf4-28ff-4d8f-8e25-737e57dcd868",
   "metadata": {
    "panel-layout": {
     "height": 0,
     "visible": true,
     "width": 100
    }
   },
   "outputs": [],
   "source": [
    "# Save result\n",
    "wb.save(\"../output/final_lab5V2.xlsx\")"
   ]
  }
 ],
 "metadata": {
  "kernelspec": {
   "display_name": "Python 3 (ipykernel)",
   "language": "python",
   "name": "python3"
  },
  "language_info": {
   "codemirror_mode": {
    "name": "ipython",
    "version": 3
   },
   "file_extension": ".py",
   "mimetype": "text/x-python",
   "name": "python",
   "nbconvert_exporter": "python",
   "pygments_lexer": "ipython3",
   "version": "3.13.5"
  },
  "panel-cell-order": [
   "51dad2ae-e5be-4884-bf2b-5732eb71d4fe",
   "b1397479-ebb3-460c-9f72-43aba5515af7"
  ]
 },
 "nbformat": 4,
 "nbformat_minor": 5
}
